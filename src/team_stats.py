# -*- coding: utf-8 -*-
from __init__ import *
import pandas as pd
from pymongo import MongoClient
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from src.utils.tools import parse_match_time

client = MongoClient('mongodb://basketball:basketball@localhost:37017/')
match_db = client['比赛']
season_db = client['赛季']
user_db = client['球员']

# 团队统计指标
team_stats_titles = [
    '得分',
    '篮板',
    '前场板',
    '后场板',
    '助攻',
    '盖帽',
    '抢断',
    '失误',
    '投篮命中',
    '投篮出手',
    '3分命中',
    '3分出手',
    '罚球命中',
    '罚球出手',
]


# 计算团队统计数据
def calculate_team_stats():
    """计算所有比赛的团队统计数据"""
    # 获取所有比赛
    match_names = match_db.list_collection_names()
    sorted_matches = sorted(
        match_names,
        key=lambda x: parse_match_time(x) if parse_match_time(x) else '',
    )

    # 存储所有比赛的团队统计
    all_team_stats = []

    for match_name in sorted_matches:
        # 获取比赛数据
        match_data = list(match_db[match_name].find())
        if not match_data:
            continue

        # 分离两队数据
        teams = {}
        for player in match_data:
            team_name = player.get('球队', '未知')
            if team_name not in teams:
                teams[team_name] = []
            teams[team_name].append(player)

        # 计算每队的统计数据
        for team_name, players in teams.items():
            # 基本信息
            match_time = parse_match_time(match_name)
            win_status = '胜' if players and players[0].get(
                '胜负') == '胜' else '负'

            # 初始化团队统计
            team_stats = {
                '比赛': match_name,
                '比赛时间': match_time,
                '球队': team_name,
                '胜负': win_status,
                '人数': len(players)
            }

            # 计算团队总数据
            for stat in team_stats_titles:
                team_stats[stat] = sum(
                    player.get(stat, 0) for player in players)

            # 计算投篮命中率
            if team_stats['投篮出手'] > 0:
                team_stats['投篮命中率'] = round(
                    team_stats['投篮命中'] / team_stats['投篮出手'], 3)
            else:
                team_stats['投篮命中率'] = 0

            # 计算三分命中率
            if team_stats['3分出手'] > 0:
                team_stats['3分命中率'] = round(
                    team_stats['3分命中'] / team_stats['3分出手'], 3)
            else:
                team_stats['3分命中率'] = 0

            # 计算罚球命中率
            if team_stats['罚球出手'] > 0:
                team_stats['罚球命中率'] = round(
                    team_stats['罚球命中'] / team_stats['罚球出手'], 3)
            else:
                team_stats['罚球命中率'] = 0

            # 添加到结果列表
            all_team_stats.append(team_stats)

    return all_team_stats


# 生成团队统计Excel
def generate_team_stats(output_path="output/团队统计.xlsx"):
    team_stats = calculate_team_stats()
    df = pd.DataFrame(team_stats)
    columns = [
        '比赛', '比赛时间', '球队', '胜负', '人数', '得分', '篮板', '前场板', '后场板', '助攻', '盖帽',
        '抢断', '失误', '投篮命中率', '3分命中率', '罚球命中率', '投篮命中', '投篮出手', '3分命中', '3分出手',
        '罚球命中', '罚球出手'
    ]

    # 确保所有列都存在
    for col in columns:
        if col not in df.columns:
            df[col] = 0

    # 按比赛时间排序
    df = df.sort_values(by=['比赛时间'], ascending=False)

    # 保存到Excel
    with pd.ExcelWriter(output_path) as writer:
        df.to_excel(writer, sheet_name='团队统计', index=False, columns=columns)

    format_team_stats_excel(output_path, columns)

    print(f"团队统计已保存到 {output_path}")
    return output_path


# 格式化Excel
def format_team_stats_excel(file_path, columns):
    wb = load_workbook(file_path)
    ws = wb['团队统计']

    # 设置列宽
    ws.column_dimensions['A'].width = 15  # 比赛
    ws.column_dimensions['B'].width = 12  # 比赛时间
    ws.column_dimensions['C'].width = 10  # 球队
    ws.column_dimensions['D'].width = 6  # 胜负
    ws.column_dimensions['E'].width = 6  # 人数

    # 设置数字列的宽度
    for i in range(5, len(ws.column_dimensions) + 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 10

    # 设置标题行样式
    header_fill = PatternFill(start_color="DDEBF7",
                              end_color="DDEBF7",
                              fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 设置数据行居中对齐
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # 添加条件格式
    score_col = get_column_letter(columns.index('得分') + 1)
    ws.conditional_formatting.add(
        f"{score_col}2:{score_col}{ws.max_row}",
        ColorScaleRule(start_type='min',
                       start_color='FFFFFF',
                       end_type='max',
                       end_color='FF8C00'))

    # 命中率列
    for stat in ['投篮命中率', '3分命中率', '罚球命中率']:
        col = get_column_letter(columns.index(stat) + 1)
        ws.conditional_formatting.add(
            f"{col}2:{col}{ws.max_row}",
            ColorScaleRule(start_type='min',
                           start_color='FFFFFF',
                           end_type='max',
                           end_color='4CAF50'))

    # 保存格式化后的文件
    wb.save(file_path)


# 直接运行时执行
if __name__ == '__main__':
    generate_team_stats("output/团队统计.xlsx")
