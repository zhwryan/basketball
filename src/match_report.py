# -*- coding: utf-8 -*-

from __init__ import *
import json
from ntpath import isfile
from openpyxl import load_workbook
from pymongo import MongoClient
from src.utils.open_api import OpenApi
from src.utils.utils import Utils

client = MongoClient('mongodb://basketball:basketball@localhost:37017/')
player_db = client['球员']
season_db = client['赛季']
match_db = client['比赛']

_SYSTEM = """
- Role: 专业篮球评论员
- Background: 用户需要一份激情四溢的比赛报告，这份报告要基于球员和球队的的数据，展现比赛精彩，让读者感受到比赛的紧张与激情。
- Profile: 你是一位资深的篮球评论员，对篮球比赛有着敏锐的洞察力和丰富的解说经验，能够迅速从海量数据中提炼出关键信息，并用生动的语言将其转化为引人入胜的比赛故事。
- Skills: 你具备出色的篮球知识、数据分析能力、语言表达能力和故事叙述技巧，能够将枯燥的数据转化为充满激情和感染力的文字，让读者感受到比赛的每一个精彩瞬间。
- Goals: 根据提供的球员和球队数据，撰写一份激情澎湃的比赛报告，突出球队的数据、球员的数据，让读者感受到比赛的紧张与激情。
- Constrains:
  - 保持客观公正的态度，确保数据的准确性和分析的合理性.
  - 杜绝编写比赛过程,严格从数据角度呈现,不要编造比赛过程.
- OutputFormat: 比赛报告应包括比赛概述、球队数据对比、关键球员表现和比赛总结四个部分，语言风格应充满激情和感染力，能够吸引读者的注意力并激发他们的兴趣。
- Workflow:
  1. 梳理比赛数据，提取关键信息。
  2. 对比两队的数据差异,分析两队的数据，总结比赛的胜负原因.
  3. 对球员表现进行点评，分析对他们的数据，总结比赛的胜负原因。
  4. 总结比赛胜负关键,建议后续改进计划.
  5. 检查报告中的数据是否正确,避免数据错误.
- Examples:
  - 比赛概述
  - 球队数据对比
  - 关键球员表现
  - 比赛总结
"""


def query_team_stats(match_name):
    """
    从Excel文件中查询球队统计数据
    :param match_name: 比赛名称
    :return: 格式化后的球队统计数据字符串列表,每条记录一个JSON
    """
    wb = load_workbook('output/团队统计.xlsx')
    sheet = wb.active
    teams = []
    headers = [cell.value for cell in sheet[1]]  # 获取表头
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始遍历数据
        if row[0] == match_name:  # 第一列是比赛名称
            col = dict(zip(headers, row))
            teams.append(json.dumps(col, ensure_ascii=False))

    return "\n".join(teams)


def query_players_stats(match_name):
    """
    查询球员统计数据
    :param sheet_name: 工作表名称，格式如"25.4.25白胜"
    :return: 格式化后的球员统计数据字符串
    """
    result = []
    for match_data in match_db[match_name].find():
        match_data.pop('投票')
        name = match_data.pop("_id")
        avg_data = season_db['场均数据'].find_one({"姓名": name})
        if avg_data:
            avg_data.pop('_id')
            avg_data.pop('场次')
        pos_data = player_db['名单'].find_one({"姓名": name})
        if pos_data:
            pos_data.pop('_id')
        result.append(f"""
## {name}球员数据
```json
// 本场比赛数据
{json.dumps(match_data, ensure_ascii=False)}
// 历史场均数据
{json.dumps(avg_data or {}, ensure_ascii=False)}
// 球员个人信息
{json.dumps(pos_data or {}, ensure_ascii=False)}
```""")

    return "\n".join(result)


def query_match_event(match_name):
    result = []
    src_dir = f"userdata/{match_name}"
    for file in os.listdir(src_dir):
        path = op.join(src_dir, file)
        if not isfile(path):
            continue
        with open(path, 'r', encoding='utf-8') as f:
            result.append(f.read())
    return "\n".join(result)


def gen_prompt(match_name, scores):
    return f"""
# 球队数据
```json
{query_team_stats(match_name)}
```

# 四节比分
```
{scores}
```

# 球员数据
{query_players_stats(match_name)}

# 比赛过程
```
{query_match_event(match_name)}
```
"""


def main(match_name, scores, use_llm=True):
    prompt = gen_prompt(match_name, scores)

    if use_llm:
        api = OpenApi("deepseek")
        api.system = _SYSTEM
        report = api.generate(prompt, temperature=0.0)
        Utils.write_file('output/比赛报告.md', report)
    else:
        print(prompt)


# 示例用法
if __name__ == '__main__':
    scores = "白队 19 40 32 29\n紫队 30 31 19 37"
    match_name = "25.4.25白胜"
    main(
        match_name,
        scores,
        # False,
    )
