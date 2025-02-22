# -*- coding: utf-8 -*-

import pandas as pd
from pymongo import MongoClient
from tools import *
import openpyxl
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

match_titles = [
    '得分', '篮板', '前场板', '后场板', '助攻', '盖帽', '抢断', '失误', '投篮', '3分', '罚球', '效率值',
    '攻防贡献率', '百回合得分', '球队', '胜负', '投票'
]
score_titles = ['胜场', '败场', '积分', '胜率', '连胜', '连败']
avg_titles = [
    '得分', '篮板', '前场板', '后场板', '助攻', '盖帽', '抢断', '失误', '效率值', '投篮命中', '投篮出手',
    '投篮命中率', '3分命中', '3分出手', '3分命中率', '罚球命中', '罚球出手', '罚球命中率', '有效命中率', '真实命中率'
]
season_titles = [
    '得分', '篮板', '助攻', '盖帽', '抢断', '失误', '前场板', '后场板', '投篮命中', '投篮出手', '3分命中',
    '3分出手', '罚球命中', '罚球出手'
]
client = MongoClient('mongodb://basketball:basketball@localhost:37017/')
match_db = client['比赛']
season_db = client['赛季']
user_db = client["球员"]


# 生成比赛数据
def generate_match_db(src_path):
    clear_all_matches()

    file = pd.ExcelFile(src_path)
    for sheet_name in file.sheet_names:
        if sheet_name in match_db.list_collection_names():
            continue

        sheet = pd.DataFrame(file.parse(sheet_name))
        match_time = parse_match_time(sheet_name)
        users = []
        for record in sheet.to_dict(orient='records'):
            name = record['姓名']
            u = {'_id': name, '姓名': name}
            for title in record.keys():
                if title in match_titles:
                    u[title] = record[title]
                if title in ['投篮', '3分', '罚球']:
                    u[title + '命中'] = int(record[title].split('-')[0])
                    u[title + '出手'] = int(record[title].split('-')[1])

            # 计算进攻回合数
            possessions = (
                u['投篮出手'] + u['3分出手']  # 投篮出手数
                + u['失误']  # 失误数
                + 0.44 * u['罚球出手']  # 罚球系数
                - u['前场板'])  # 减去进攻篮板

            # 百回合得分
            if possessions > 0:
                u['百回合得分'] = round(u['得分'] / possessions * 100, 1)
            else:
                u['百回合得分'] = 0

            u['效率值'] = calc_per(u, 1)
            u['比赛时间'] = match_time
            users.append(u)

        # 全队贡献度
        total_ows = sum(calc_ows(user) for user in users)
        total_dws = sum(calc_dws(user) for user in users)
        for u in users:
            u['攻防贡献率'] = calc_ws(u, total_ows, total_dws)
            u['投票'] = f"{u['姓名']} {u['得分']}分{u['篮板']}板{u['助攻']}助"
            u['投票'] += f" 投篮{ u['投篮'].replace('-', '/') }"
            u['投票'] += f" 效率{u['效率值']} 贡献{round(u['攻防贡献率']*100, 2)}%"
            if '胜' in u['胜负']:
                u['投票'] += " 喜胜"
            else:
                u['投票'] += " 惜败"
        update_collection(match_db, sheet_name, users)


# 生成赛季数据
def generate_season_db():
    users = {}
    for match_name in match_db.list_collection_names():
        for match_data in match_db[match_name].find():
            name = match_data['姓名']
            if not user_db["名单"].find_one({'_id': name}):
                continue
            u = users.get(name, {'_id': name, '场次': 0, '姓名': name})
            u['场次'] += 1
            for title in season_titles:
                u[title] = u.get(title, 0) + match_data.get(title, 0)
            users[name] = u

    update_collection(season_db, '赛季数据', list(users.values()))


# 生成场均数据
def generate_avg_db():
    records = {}
    for season in season_db['赛季数据'].find():
        name = season['姓名']
        times = season['场次']
        user = {
            '_id': name,
            '姓名': name,
            '场次': times,
            '真实命中率': calc_ts(season),
            '有效命中率': calc_efg(season),
            '投篮命中率': calc_fg(season, "投篮"),
            '3分命中率': calc_fg(season, "3分"),
            '罚球命中率': calc_fg(season, "罚球"),
            '效率值': calc_per(season, times),
        }
        for title in avg_titles:
            if title in season:
                user[title] = round(season[title] / times, 2)
        records[name] = user

    update_collection(season_db, '场均数据', list(records.values()))


# 生成积分榜
def generate_score_db():
    match_names = match_db.list_collection_names()
    sorted_matches = sorted(
        match_names,
        key=lambda x: parse_match_time(x),
    )

    scores = {}
    for user in season_db['赛季数据'].find():
        name = user['姓名']
        score = {title: 0 for title in score_titles}
        score['_id'] = name
        score['姓名'] = name
        score['场次'] = user['场次']

        for match_name in sorted_matches:
            match_user = match_db[match_name].find_one({'_id': name})
            if not match_user:
                continue
            if match_user['胜负'] and match_user['胜负'] == '胜':
                score['胜场'] += 1
                score['连胜'] += 1
                score['连败'] = 0
            else:
                score['败场'] += 1
                score['连胜'] = 0
                score['连败'] += 1
            score['积分'] = score['胜场'] * 2 + score['败场']
            score['胜率'] = round(score['胜场'] / score['场次'], 2)
        scores[name] = score

    update_collection(season_db, '积分榜', list(scores.values()))


# 生成excel
def generate_excel(path):
    sorters = {
        '积分榜': score_titles,
        '场均数据': avg_titles,
        '赛季数据': season_titles,
    }
    with pd.ExcelWriter(path) as f:
        for col in sorters:
            data = season_db[col].find()
            df = pd.DataFrame(data)
            df = df.drop(columns=['_id'])
            columns = sorters[col].copy()
            columns.insert(0, '姓名')
            columns.insert(1, '场次')
            df.to_excel(f, sheet_name=col, index=False, columns=columns)

        matches = match_db.list_collection_names()
        sorted_matches = sorted(
            matches,
            key=lambda x: parse_match_time(x),
            reverse=True,
        )
        for col in sorted_matches:
            data = match_db[col].find()
            df = pd.DataFrame(data)
            df = df.drop(columns=['_id'])
            columns = match_titles.copy()
            columns.insert(0, '姓名')
            df.to_excel(f, sheet_name=col, index=False, columns=columns)


# 更新集合
def update_collection(db, collection_name, records):
    if collection_name in db.list_collection_names():
        db.drop_collection(collection_name)
    db[collection_name].insert_many(records)


# 清理比赛集合
def clear_all_matches():
    for match_name in match_db.list_collection_names():
        match_db.drop_collection(match_name)


def get_column_numeric_values(worksheet, col_letter):
    """获取指定列的数值数据"""
    data_range = f"{col_letter}2:{col_letter}{worksheet.max_row}"
    values = []
    for row in worksheet[data_range]:
        for cell in row:
            if cell.value is not None:
                try:
                    num_value = float(cell.value)
                    values.append(num_value)
                except (ValueError, TypeError):
                    continue
    return values


# 格式化列
def format_column(worksheet, column):
    col_letter = get_column_letter(column)

    # 为单个列应用色阶规则
    color_scale_rule = ColorScaleRule(
        start_type='min',
        start_color='FFFFFFFF',  # 白色
        end_type='max',
        end_color='FB807B'  # 红色
    )
    apply_range = f"${col_letter}2:${col_letter}{worksheet.max_row}"
    worksheet.conditional_formatting.add(apply_range, color_scale_rule)

    # 为最后一个字是'率'的列设置百分比格式
    col_letter = get_column_letter(column)
    header_cell = worksheet[f"{col_letter}1"]
    if header_cell.value and isinstance(
            header_cell.value, str) and header_cell.value.endswith("率"):
        for cell in worksheet[col_letter][1:]:  # 跳过表头
            if cell.value is not None:
                try:
                    cell.number_format = '0.00%'
                except (ValueError, TypeError):
                    continue

    # 设置最佳列宽
    max_length = 0
    for cell in worksheet[col_letter]:
        if cell.value:
            # 计算字符串实际显示宽度（中文字符计为2，其他字符计为1）
            cell_length = sum(2 if '\u4e00' <= char <= '\u9fff' else 1
                              for char in str(cell.value))
            max_length = max(max_length, cell_length)
    adjusted_width = max_length + 4  # 额外添加4个字符宽度作为缓冲
    worksheet.column_dimensions[col_letter].width = adjusted_width


# 格式化子表
def format_sheet(worksheet):
    # 设置第一行高度为30
    worksheet.row_dimensions[1].height = 30

    # 设置第一行文字置顶对齐
    for cell in worksheet[1]:
        cell.alignment = Alignment(vertical='top')

    # 添加数据筛选
    max_col = worksheet.max_column
    max_row = worksheet.max_row
    worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    # 冻结首行首列
    worksheet.freeze_panes = 'B2'


# 格式化表格
def formal_excel(src_path):
    workbook = openpyxl.load_workbook(src_path)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        format_sheet(worksheet)
        for column in range(2, worksheet.max_column + 1):
            format_column(worksheet, column)

    workbook.save(src_path)


if __name__ == '__main__':
    src_path = 'assets/数据统计.xlsx'
    out_path = "output/赛季数据.xlsx"
    generate_match_db(src_path)
    generate_season_db()
    generate_avg_db()
    generate_score_db()
    generate_excel(out_path)
    formal_excel(out_path)
